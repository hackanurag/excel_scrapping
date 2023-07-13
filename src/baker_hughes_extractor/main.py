from sqlalchemy import create_engine
from sqlalchemy.engine import URL
from sqlalchemy import text
from collections import Counter
import openpyxl
import sys
import os

#connection details to my local postgres server
hostx = "127.0.0.1"
userx = "<user_name>"
passwordx = "<user_password>"


class style():
    BLACK = '\033[30m'
    RED = '\033[31m'
    GREEN = '\033[32m'
    YELLOW = '\033[33m'
    BLUE = '\033[34m'
    MAGENTA = '\033[35m'
    CYAN = '\033[36m'
    WHITE = '\033[37m'
    UNDERLINE = '\033[4m'
    RESET = '\033[0m'

#function to get column list based on range
def column_listx(x):
    column_list = []
    init_list = []
    prefix = []
    final_list = []
    if x < 26:
        for i in range(x):
            var = 65 + i
            init_list.append(chr(var))
            column_list.append(chr(var))
    elif x > 26:
        for i in range(26):
            var = 65 + i
            init_list.append(chr(var))
            column_list.append(chr(var))
        quot = int(x / 26)

        for j in range(quot):
            prefix.append(chr(65 + j))
        for k in prefix:
            for l in init_list:
                column_list.append(k + l)
    for i in range(x):
        final_list.append(column_list[i])
    return final_list

#function to scrape data and create map
def etl(sheet):
    state_list = []
    land_offshore_map = {"2": {"Date": [], "Land": [], "Offshore": []}, "1": {"Date": [], "Land": []},
                         "3": {"Date": [], "Land": [], "Offshore": [], "Total": []}}
    for row in sheet.iter_rows(min_row=6, max_row=6):
        for cell in row:
            state_list.append(str(cell.value))

    column_list = column_listx(len(state_list))

    state_list.pop(0)
    tempo = []
    for i in state_list:
        tempo.append(i)
    for state in range(len(state_list)):
        if state_list[state] == 'None':
            tempo[state] = state_list[state - 1]
    count_state_occurrence = dict(Counter(tempo))
    for i in count_state_occurrence.keys():
        if "TOTAL" in i:
            count_state_occurrence[i] = int(count_state_occurrence[i]) + 1

    for i in count_state_occurrence.keys():
        count_state_occurrence[i] = land_offshore_map[str(count_state_occurrence[i])]

    del count_state_occurrence['None']
    map_level_0 = dict(count_state_occurrence)

    state_list_solid = []
    for i in state_list:
        if i != 'None':
            state_list_solid.append(i)
    level_1 = []

    for i in column_list:
        tempv = []

        for x in range(7, len(sheet[f'{i}'])):
            tempv.append(sheet[f'{i}'][x].value)
        level_1.append(tempv)

    datex = level_1[0]

    temp = []
    for i in datex:
        cc = str(i.strftime("%m/%d/%Y"))
        xx = f"'{cc}'"
        temp.append(xx)
    datex = temp

    map_final = {}

    counter = 1
    for i in state_list_solid:
        tempc = {}
        for j in map_level_0[i].keys():
            try:
                if j != 'Date':
                    tempc[j] = level_1[counter]
                    counter += 1

            except Exception as e:
                print(e)
            map_final[i] = tempc

    for i in state_list_solid:
        map_final[i]['Date'] = datex

    return map_final


url = URL.create(
    drivername="postgresql",
    username=userx,
    password=passwordx,
    host=hostx,
    database="postgres"
)

engine = create_engine(url)
conn = engine.connect()
conn.execute(text("commit"))
conn.execute(text("create database <database_name>;"))
conn.close()

#function to create and insert tables into Postgres using map created
def data_push(country_map):
    url = URL.create(
        drivername="postgresql",
        username=userx,
        password=passwordx,
        host=hostx,
        database="<database_name>"
    )
    engine = create_engine(url)
    conn = engine.connect()
    conn.execute(text("commit"))
    for i in country_map.keys():
        column_stmt = ''
        column_stmt_for_insert = ''

        for j in country_map[i].keys():
            if j != 'Date':
                column_stmt += j + ' int, '
                column_stmt_for_insert += j + ', '
        column_stmt += 'Date varchar(255) PRIMARY KEY'
        column_stmt_for_insert += 'Date'
        formating_only = '"' + i + '"'
        sql_statement_to_create_table = f"CREATE TABLE {formating_only} ({column_stmt}) ;"
        print(style.YELLOW + f"Creating table {formating_only}...")
        conn.execute(text(sql_statement_to_create_table))
        temp = []
        for j in country_map[i].keys():
            temp.append(country_map[i][j])

        for l in range(len(temp[0])):
            values = ''
            for k in temp:
                values += str(k[l]) + ', '
            values = values[:len(values) - 2]
            sql_statement_to_insert = f"INSERT INTO {formating_only} ({column_stmt_for_insert}) VALUES ({values});"
            print(style.YELLOW + f"Inserting data {values} into table {formating_only}...")
            conn.execute(text(sql_statement_to_insert))

#relative path so that can work anywhere
path = os.path.dirname(os.path.abspath(sys.argv[0]))
rel_path = path.replace("src/baker_hughes_extractor", "")
wb = openpyxl.load_workbook(f'{rel_path}bh_ data.xlsx')

for i in wb.sheetnames:
    print(style.GREEN + f"Started scraping and ETL process for {str(wb[i])}...\n\n")
    countryx = etl(wb[i])

    print(style.GREEN + f"Started pushing {str(wb[i])} data to POSTGRES-SERVER...\n\n")
    data_push(countryx)

conn.close()

print(style.GREEN + "DATA PUSH DONE!")
